"""
Excel / URL I/O — records destruction events and loads tank lists from Excel or URLs.
"""

import sys
import traceback
import urllib.request
import urllib.parse
from pathlib import Path

from .constants import EXCEL_LOCK

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ── helpers ────────────────────────────────────────────────────────────────────

def _is_url(path: str) -> bool:
    return path.startswith("http://") or path.startswith("https://")


def _normalize_excel_url(url: str) -> str:
    """
    Converts Google Sheets / Google Drive share URLs to a direct CSV export URL.
    Also passes through OneDrive and other direct links unchanged.
    """
    if "docs.google.com/spreadsheets" in url:
        import re
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
        if m:
            sheet_id = m.group(1)
            gid_m = re.search(r"[#&?]gid=(\d+)", url)
            gid = gid_m.group(1) if gid_m else "0"
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    if "drive.google.com/file/d/" in url:
        import re
        m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
        if m:
            return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    return url


# ── destruction recording ──────────────────────────────────────────────────────

def _normalize_tank_name(name: str) -> str:
    if not name: return ""
    import unicodedata
    s = "".join(c for c in unicodedata.normalize('NFD', str(name)) if unicodedata.category(c) != 'Mn')
    s = s.lower()
    for char in " .-/_'\"":
        s = s.replace(char, "")
    s = s.replace("object", "obj").replace("pzkpfw", "pz").replace("panhard", "")
    return s

def record_destruction(
    path: str,
    player: str,
    tank: str,
    map_name: str,
    battle_time: str,
    all_players: list[str] = None,
):
    if not path:
        return

    # Clean player name (remove [TAG] etc.)
    if " " in player:
        player = player.split(" ")[-1]
    if "]" in player:
        player = player.split("]")[-1]
    player = player.strip()

    if _is_url(path):
        try:
            params_dict = {
                "player": player,
                "tank":   tank,
                "status": "Destroyed",
            }
            params = urllib.parse.urlencode(params_dict)
            url = f"{path}?{params}"
            with urllib.request.urlopen(url, timeout=10) as resp:
                resp.read()
        except Exception as e:
            print(f"[!] Results Export Error: {e}", file=sys.stderr)
    else:
        if not HAS_OPENPYXL:
            return

        with EXCEL_LOCK:
            try:
                from openpyxl import Workbook, load_workbook
                p = Path(path)
                if p.exists():
                    wb = load_workbook(path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    wb.save(path)

                # 1. Search ALL columns in row 1 for the player name
                target_col = -1
                last_used_col = 0
                for c in range(1, 1001):
                    val = ws.cell(row=1, column=c).value
                    if val:
                        last_used_col = c
                        if str(val).strip().lower() == player.lower():
                            target_col = c
                            break

                # 2. If not found, find the spot for a new player
                if target_col == -1:
                    target_col = last_used_col + 2 if last_used_col > 0 else 1
                    ws.cell(row=1, column=target_col, value=player)

                if target_col != -1:
                    # 3. Find tank in this player's list or find first empty row
                    row = 2
                    found_row = -1
                    tank_norm = _normalize_tank_name(tank)
                    while True:
                        t_val = ws.cell(row=row, column=target_col).value
                        if not t_val:
                            found_row = row
                            break
                        
                        t_val_norm = _normalize_tank_name(str(t_val))
                        if t_val_norm == tank_norm or (t_val_norm and tank_norm and (t_val_norm in tank_norm or tank_norm in t_val_norm)):
                            found_row = row
                            break
                        row += 1

                    # 4. Update status with counter
                    ws.cell(row=found_row, column=target_col, value=tank)

                    status_cell = ws.cell(row=found_row, column=target_col + 1)
                    current_val = str(status_cell.value or "").strip()

                    count = 1
                    if "X" in current_val.upper():
                        try:
                            parts = current_val.upper().split("X")
                            if parts[-1].isdigit():
                                count = int(parts[-1]) + 1
                        except Exception:
                            count = 1
                    elif current_val.lower() == "destroyed":
                        count = 2

                    status_cell.value = f"Destroyed X{count}"

                wb.save(path)
            except Exception as e:
                print(f"[!] Failed to record destruction to Excel: {e}", file=sys.stderr)


# ── tank list loading ──────────────────────────────────────────────────────────

def load_tanks_from_excel(
    path: str,
    clan_members: set[str] | None = None,
) -> dict[str, list[str] | str]:
    """
    Reads tank names from columns of an Excel file or URL.
    Assigns each column to a player (first non-numeric row is player name, rest are tanks).
    """
    result: dict[str, list[str] | str] = {}
    normalized_members = {m.lower() for m in clan_members} if clan_members is not None else None

    def _process_cells(cells):
        name = None
        tanks = []
        for cell in cells:
            s_cell = str(cell).strip()
            if not s_cell:
                continue
            if not name:
                # Skip if it looks like a number (tier/count)
                try:
                    float_val = float(s_cell.replace(",", "."))
                    if float_val < 1000: # Heuristic: account IDs are usually larger
                        continue
                except ValueError:
                    pass
                
                name = s_cell.replace("⭐", "").strip()
            else:
                t = s_cell.replace("⭐", "").strip()
                if t:
                    tanks.append(t)
        
        if name:
            # Only filter if clan_members was explicitly provided (even if empty)
            if clan_members is not None:
                if name.lower() not in normalized_members:
                    return
            result[name.lower()] = tanks
            result[f"{name.lower()}_display"] = name

    if _is_url(path):
        if not HAS_PANDAS:
            return {}
        try:
            url = _normalize_excel_url(path)
            try:
                df = pd.read_csv(url, header=None)
            except Exception:
                df = pd.read_excel(url, header=None)
            for col_idx in range(df.shape[1]):
                col = df.iloc[:, col_idx].dropna().astype(str).str.strip()
                col = col[col != ""]
                if not col.empty:
                    _process_cells(list(col))
            return result
        except Exception as e:
            print(
                f"[!] Failed to load tank list from URL: {e}\n{traceback.format_exc()}",
                file=sys.stderr,
            )
            return {}
    else:
        if not HAS_OPENPYXL:
            print("[!] openpyxl not installed — cannot read local Excel files", file=sys.stderr)
            return {}
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for col in ws.iter_cols(values_only=True):
                cells = [str(x).strip() for x in col if x is not None and str(x).strip()]
                if cells:
                    _process_cells(cells)
            wb.close()
            return result
        except Exception as e:
            print(f"[!] Failed to load Excel file: {e}", file=sys.stderr)
            return {}
